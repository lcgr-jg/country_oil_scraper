"""One-off probe: BAFA Mineralöldaten format eras (xlsx vs pdf)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pdfplumber
import requests
from bs4 import BeautifulSoup

RAW = Path("data/raw/germany/probe")
RAW.mkdir(parents=True, exist_ok=True)


def sniff_and_save(month_key: str) -> Path | None:
    url = (
        "https://www.bafa.de/SharedDocs/Downloads/DE/Energie/Mineraloel/"
        f"moel_amtliche_daten_{month_key}.pdf?__blob=publicationFile&v=2"
    )
    r = requests.get(url, timeout=60)
    if r.status_code != 200:
        print(month_key, "HTTP", r.status_code)
        return None
    if r.content[:4] == b"%PDF":
        dest = RAW / f"moel_amtliche_daten_{month_key}.pdf"
    elif r.content[:2] == b"PK":
        # URL says .pdf but payload is Office Open XML (xlsx)
        dest = RAW / f"moel_amtliche_daten_{month_key}.xlsx"
    else:
        print(month_key, "unknown magic", r.content[:20])
        return None
    dest.write_bytes(r.content)
    print(month_key, "->", dest.name, dest.stat().st_size, r.headers.get("content-type"))
    return dest


def inspect_xlsx(path: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    print("==", path.name, "sheets", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f" sheet {sn!r} max_row={ws.max_row} max_col={ws.max_column}")
        shown = 0
        for r in range(1, min(ws.max_row + 1, 40)):
            vals = [ws.cell(r, c).value for c in range(1, min(10, ws.max_column + 1))]
            if any(v is not None and str(v).strip() for v in vals):
                print(" ", r, vals)
                shown += 1
                if shown >= 18:
                    break


def landing_downloads(slug: str) -> None:
    url = (
        "https://www.bafa.de/SharedDocs/Downloads/DE/Energie/Mineraloel/"
        f"{slug}.html"
    )
    r = requests.get(url, timeout=60)
    print("LANDING", slug, r.status_code)
    if r.status_code != 200:
        return
    soup = BeautifulSoup(r.text, "html.parser")
    for a in soup.find_all("a", href=True):
        h = a["href"]
        if "__blob" in h or any(ext in h.lower() for ext in (".pdf", ".xlsx", ".xls", ".docx")):
            print(" ", " ".join((a.get_text() or "").split())[:70], "->", h)


def inspect_pdf(path: Path) -> None:
    with pdfplumber.open(path) as pdf:
        print("==", path.name, "pages", len(pdf.pages))
        # Decode cid fonts? try char-level
        page0 = pdf.pages[0]
        chars = page0.chars[:30] if page0.chars else []
        print(" sample chars", [(c.get("text"), c.get("fontname")) for c in chars[:15]])
        for i in [0, 1, 7, 8, 9, 10, 15, 20]:
            if i >= len(pdf.pages):
                continue
            p = pdf.pages[i]
            text = p.extract_text() or ""
            tables = p.extract_tables() or []
            print(f" page {i+1}: chars={len(text)} tables={len(tables)}")
            if tables:
                print("  table0 first rows:")
                for row in tables[0][:6]:
                    print("   ", row)
            # printable ascii-ish lines
            lines = [ln for ln in text.splitlines() if any(ch.isalpha() for ch in ln)]
            for ln in lines[:8]:
                print("  |", ln[:100])


if __name__ == "__main__":
    for m in ["2020_06", "2022_01", "2023_06", "2024_01", "2024_06", "2024_12", "2025_01"]:
        sniff_and_save(m)

    for path in sorted(RAW.glob("*.xlsx")):
        inspect_xlsx(path)

    for slug in [
        "moel_amtliche_daten_2019_dezember",
        "moel_amtliche_daten_2020_juni",
        "moel_amtliche_daten_2020_06",
        "moel_amtliche_daten_2024_06",
    ]:
        landing_downloads(slug)

    for path in sorted(RAW.glob("*.pdf")):
        if "2024_06" in path.name or "2025_01" in path.name:
            inspect_pdf(path)
