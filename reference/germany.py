"""
reference.germany
─────────────────
BAFA (Bundesamt für Wirtschaft und Ausfuhrkontrolle) — Amtliche Mineralöldaten.

Monthly domestic deliveries (Inlandsablieferungen), closing stocks
(Eigentumsendbestand), and bio blending (Beimischung). Source files are
XLSX or PDF (payload often mislabeled); see download sniffing below.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import pandas as pd
import pdfplumber
import requests

from reference.jodi_compare import JodiCompareSeries, sum_natives_series_for_jodi

logger = logging.getLogger(__name__)

BAFA_AGENCY_SOURCE = "BAFA"
BAFA_DATASET_SOURCE = "germany_bafa_mineraloel"
BAFA_UNIT_NATIVE = "t"
BAFA_DEMAND_METRIC = "TOTDEMO"
BAFA_STOCKS_METRIC = "CLOSTLV"
BAFA_BIO_METRIC = "BIOBLEND"

COUNTRY_CODE = "DE"
COUNTRY_NAME = "Germany"
SOURCE_ID = BAFA_DATASET_SOURCE
JODI_REF_AREA = "DE"

BAFA_BASE = "https://www.bafa.de"
DOWNLOAD_DIR = f"{BAFA_BASE}/SharedDocs/Downloads/DE/Energie/Mineraloel"

# Bootstrap history start (Infothek coverage from exploration).
HISTORY_START = "2019-12"

CANONICAL_COLUMNS: list[str] = [
    "date",
    "country",
    "country_name",
    "source",
    "metric_type",
    "product_native",
    "product",
    "value",
    "unit",
    "is_provisional",
    "source_file",
    "updated_at",
]

MONTH_DE: dict[str, int] = {
    "januar": 1,
    "februar": 2,
    "feburar": 2,
    "maerz": 3,
    "märz": 3,
    "marz": 3,
    "april": 4,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "dezember": 12,
}

MONTH_DE_BY_NUM = {
    v: k
    for k, v in MONTH_DE.items()
    if k not in {"feburar", "märz", "marz"}
}
MONTH_DE_BY_NUM[3] = "maerz"

_CID_BYTE_FIXUPS = str.maketrans(
    {
        "\x99": "ö",
        "\x9e": "ü",
        "\x8a": "ä",
        "\x89": "ä",
    }
)
_CID_LABEL_FIXUPS = {
    "Heiz™l": "Heizöl",
    "Roh™l": "Rohöl",
    "Mineral™l": "Mineralöl",
    "ausl‰ndisches": "ausländisches",
    "Rckst‰nde": "Rückstände",
    "Doppelz‰hlung": "Doppelzählung",
    "Leuchtle": "Leuchtöle",
    "fžr": "für",
    "k™nnen": "können",
    "Bioheizl": "Bioheizöl",
}

_PRODUCT_ALIASES: dict[str, str] = {
    "Flugturbinenkraftstoff, leicht": "Flugturb.Kraftst.,leicht",
    "Flugturbinenkraftstoff, schwer": "Flugturb.Kraftst.,schwer",
    "Flugturb.Kraftst.,schwer und andere Leuchtöle": "Flugturb.Kraftst.,schwer",
    "Schmiermittel": "Schmierstoffe",
    "Spezialbenzine": "Spezialbenzin",
    "Wachse, Paraffine, Vaseline": "Wachse,Paraffine,Vaseline",
    "Beimischung Bioethanol": "Bioethanol",
    "davon: Anteil Bioethanol an ETBE a)": "Bioethanol an ETBE a)",
    "davon: Beimischung Biodiesel (FAME), HVO": "Biodiesel (FAME), HVO, BTL",
    "Biodiesel (FAME), HVO, BTL und andere": "Biodiesel (FAME), HVO, BTL",
    "Biodiesel (FAME), HVO, BTL, Bioheizöl": "Biodiesel (FAME), HVO, BTL",
    "Motorenbenzin": "Ottokraftstoff",
    "Motorenbenzin *": "Ottokraftstoff",
    "Rohöl gesamt": "Rohöl",
    "Mineralölprodukte gesamt": "Mineralölprodukte",
    "deutsches": "Rohöl | deutsches",
    "ausländisches": "Rohöl | ausländisches",
    "Industrie-Getriebeöle": "IndustrieGetriebeöle",
    "Turbinen-, Kompressoren- und Elektroisolieröle": (
        "Turbinen, Kompressoren und Elektroisolieröle"
    ),
}

# Rollup / junk rows — keep out of the stored parquet.
_EXCLUDE_NATIVES: frozenset[str] = frozenset(
    {
        "Inlandsabsatz",
        "Gesamt",
        "Doppelzählung aus Recycling",
        "- Doppelzählung aus Recycling",
        "Rohöl | Zusammen",
        "Mineralölprodukte | Zusammen",
        "Zusammen",
        "Insgesamt",
    }
)

SEASONALITY_DEMAND_PRODUCTS: tuple[str, ...] = (
    "Ottokraftstoff",
    "Dieselkraftstoff",
    "Heizöl, leicht",
    "Heizöl, schwer",
    "Flüssiggas",
    "Rohbenzin",
    "Flugturb.Kraftst.,schwer",
    "Flugturb.Kraftst.,leicht",
    "Benzinkomponenten",
    "Bitumen",
    "Petrolkoks",
    "Raffineriegas",
    "Schmierstoffe",
    "Raffinerieeinsatzmaterial",
)

SEASONALITY_BIO_PRODUCTS: tuple[str, ...] = (
    "Bioethanol",
    "Bioethanol an ETBE a)",
    "Biodiesel (FAME), HVO, BTL",
    "davon HVO",
    "davon FAME",
    "Bioheizöl",
)

CHART_PRODUCTS: tuple[str, ...] = SEASONALITY_DEMAND_PRODUCTS

DISPLAY_LABELS: dict[str, str] = {
    "Ottokraftstoff": "Gasoline (Otto)",
    "Dieselkraftstoff": "Diesel",
    "Heizöl, leicht": "Light heating oil",
    "Heizöl, schwer": "Heavy fuel oil",
    "Flüssiggas": "LPG",
    "Rohbenzin": "Naphtha",
    "Flugturb.Kraftst.,schwer": "Jet fuel (heavy)",
    "Flugturb.Kraftst.,leicht": "Jet fuel (light)",
    "Benzinkomponenten": "Gasoline components",
    "Bitumen": "Bitumen",
    "Petrolkoks": "Petroleum coke",
    "Raffineriegas": "Refinery gas",
    "Schmierstoffe": "Lubricants",
    "Raffinerieeinsatzmaterial": "Refinery feedstock",
    "Bioethanol": "Bioethanol",
    "Bioethanol an ETBE a)": "Bioethanol in ETBE",
    "Biodiesel (FAME), HVO, BTL": "Biodiesel (FAME/HVO/BTL)",
    "davon HVO": "of which HVO",
    "davon FAME": "of which FAME",
    "Bioheizöl": "Bio heating oil",
}

UNITS_KIND: dict[str, str] = {
    "Ottokraftstoff": "gasoline",
    "Dieselkraftstoff": "diesel",
    "Heizöl, leicht": "diesel",
    "Heizöl, schwer": "fuel_oil",
    "Flüssiggas": "lpg",
    "Rohbenzin": "naphtha",
    "Flugturb.Kraftst.,schwer": "jet",
    "Flugturb.Kraftst.,leicht": "jet",
    "Benzinkomponenten": "gasoline",
    "Bitumen": "bitumen",
    "Petrolkoks": "other",
    "Raffineriegas": "other",
    "Schmierstoffe": "lubes",
    "Raffinerieeinsatzmaterial": "other",
    "Mitteldestillatkomponenten": "diesel",
    "HS-Komponenten": "fuel_oil",
    "Spezialbenzin": "gasoline",
    "Testbenzin": "gasoline",
    "Flugbenzin": "gasoline",
    "Andere Leuchtöle": "kerosene",
    "Andere Rückstände": "other",
    "Wachse,Paraffine,Vaseline": "other",
    "Mineralölprodukte": "other",
    "Rohöl": "crude",
    "Rohöl | deutsches": "crude",
    "Rohöl | ausländisches": "crude",
    "Bioethanol": "ethanol",
    "Bioethanol an ETBE a)": "ethanol",
    "Biodiesel (FAME), HVO, BTL": "diesel",
    "davon HVO": "diesel",
    "davon FAME": "diesel",
    "Bioheizöl": "diesel",
    "Biokerosin": "jet",
    "Motorenöle": "lubes",
    "Getriebeöle": "lubes",
    "Basisöle": "lubes",
    "Schmierfette": "lubes",
    "Sonstige": "lubes",
    "Maschinenöle": "lubes",
    "Hydrauliköle": "lubes",
    "IndustrieGetriebeöle": "lubes",
    "Turbinen, Kompressoren und Elektroisolieröle": "lubes",
    "Weitere, nicht aufgeführte Produkte": "other",
}

GASOLINE_JODI_NATIVES: frozenset[str] = frozenset({"Ottokraftstoff"})
DIESEL_JODI_NATIVES: frozenset[str] = frozenset({"Dieselkraftstoff"})
JET_JODI_NATIVES: frozenset[str] = frozenset({"Flugturb.Kraftst.,schwer"})
LPG_JODI_NATIVES: frozenset[str] = frozenset({"Flüssiggas"})
NAPHTHA_JODI_NATIVES: frozenset[str] = frozenset({"Rohbenzin"})
FUEL_OIL_JODI_NATIVES: frozenset[str] = frozenset({"Heizöl, schwer"})
JODI_COMPARE_SERIES: dict[str, JodiCompareSeries] = {
    "gasoline": JodiCompareSeries(
        "gasoline", "GASOLINE", "Gasoline", GASOLINE_JODI_NATIVES
    ),
    "diesel": JodiCompareSeries(
        "diesel", "GASDIES", "Diesel", DIESEL_JODI_NATIVES
    ),
    "jet_fuel": JodiCompareSeries(
        "jet_fuel", "JETKERO", "Jet fuel", JET_JODI_NATIVES
    ),
    "lpg": JodiCompareSeries("lpg", "LPG", "LPG", LPG_JODI_NATIVES),
    "naphtha": JodiCompareSeries(
        "naphtha", "NAPHTHA", "Naphtha", NAPHTHA_JODI_NATIVES
    ),
    "fuel_oil": JodiCompareSeries(
        "fuel_oil", "RESFUEL", "Fuel oil", FUEL_OIL_JODI_NATIVES
    ),
}

JODI_COMPARE_PANEL_ORDER: tuple[str, ...] = (
    "Gasoline",
    "Diesel",
    "Jet fuel",
    "LPG",
    "Naphtha",
    "Fuel oil",
)


@dataclass(frozen=True)
class MonthFile:
    path: Path
    kind: str  # "xlsx" | "pdf"
    year: int
    month: int
    url: str = ""


# Back-compat alias for the exploration notebook.
ProbeFile = MonthFile


def parse_german_number(raw: object) -> Optional[float]:
    """Parse BAFA cell/text numbers; return None for blanks / confidentiality."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text in {"-", "–", "—", ".", "a)", "b)", "c)", "*"}:
        return None
    if text.upper() in {"#REF!", "#N/A", "NONE"}:
        return None
    text = (
        text.replace("\u00ad", "-")
        .replace("\u2212", "-")
        .replace("\u2013", "-")
        .replace("\xa0", "")
        .replace(" ", "")
    )
    if re.fullmatch(r"[+-]?\d{1,3}(\.\d{3})+(,\d+)?", text) or re.fullmatch(
        r"[+-]?\d+,\d+", text
    ):
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def decode_bafa_cid_text(text: str, offset: int = 29) -> str:
    """Decode mid-2024 BAFA PDFs that embed Latin as ``(cid:N)`` without ToUnicode."""

    def _repl(match: re.Match[str]) -> str:
        code = int(match.group(1)) + offset
        if 0 <= code <= 0x10FFFF:
            return chr(code)
        return match.group(0)

    decoded = re.sub(r"\(cid:(\d+)\)", _repl, text).translate(_CID_BYTE_FIXUPS)
    for bad, good in _CID_LABEL_FIXUPS.items():
        decoded = decoded.replace(bad, good)
    return decoded


def normalize_product_native(label: str) -> str:
    """Collapse era/label drift so XLSX and PDF months join cleanly."""
    s = re.sub(r"\s+", " ", str(label)).strip()
    s = s.translate(_CID_BYTE_FIXUPS)
    for bad, good in _CID_LABEL_FIXUPS.items():
        s = s.replace(bad, good)
    # Soft hyphens from PDF wraps (Industrie­Getriebeöle → IndustrieGetriebeöle).
    s = s.replace("\u00ad", "")
    return _PRODUCT_ALIASES.get(s, s)


def candidate_download_urls(year: int, month: int) -> list[str]:
    """Ordered URL guesses; BAFA naming changed across eras."""
    ym = f"{year}_{month:02d}"
    ym_nopad = f"{year}_{month}"
    month_de = MONTH_DE_BY_NUM[month]
    stems = [
        f"moel_amtliche_daten_{ym}",
        f"moel_amtliche_daten_{ym_nopad}",
        f"moel_amtliche_daten_{year}_{month_de}",
    ]
    seen: set[str] = set()
    urls: list[str] = []
    for stem in stems:
        if stem in seen:
            continue
        seen.add(stem)
        for ext in ("xlsx", "pdf"):
            urls.append(f"{DOWNLOAD_DIR}/{stem}.{ext}?__blob=publicationFile&v=2")
    return urls


def download_month(
    year: int,
    month: int,
    dest_dir: Path,
    *,
    session: Optional[requests.Session] = None,
    force: bool = False,
) -> MonthFile:
    """Download one month, sniffing real payload type (XLSX often behind .pdf URL)."""
    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    sess = session or requests.Session()

    if not force:
        for kind in ("xlsx", "pdf"):
            existing = dest_dir / f"moel_amtliche_daten_{year}_{month:02d}.{kind}"
            if existing.exists() and existing.stat().st_size > 1000:
                return MonthFile(existing, kind, year, month, url="")

    last_err: Optional[Exception] = None
    for url in candidate_download_urls(year, month):
        try:
            resp = sess.get(url, timeout=60)
        except requests.RequestException as exc:
            last_err = exc
            continue
        if resp.status_code != 200:
            continue
        magic = resp.content[:4]
        if magic == b"%PDF":
            kind = "pdf"
        elif magic[:2] == b"PK":
            kind = "xlsx"
        else:
            continue
        path = dest_dir / f"moel_amtliche_daten_{year}_{month:02d}.{kind}"
        path.write_bytes(resp.content)
        logger.info("Downloaded %s (%s, %s bytes)", path.name, kind, len(resp.content))
        return MonthFile(path, kind, year, month, url=url)

    raise FileNotFoundError(
        f"No BAFA file for {year}-{month:02d} (last error: {last_err!r})"
    )


def month_grid(start: str, end: str) -> list[tuple[int, int]]:
    idx = pd.period_range(start=start, end=end, freq="M")
    return [(p.year, p.month) for p in idx]


def download_many(
    months: Iterable[tuple[int, int]],
    dest_dir: Path,
    *,
    force: bool = False,
) -> list[MonthFile]:
    out: list[MonthFile] = []
    errors: list[str] = []
    with requests.Session() as session:
        for year, month in months:
            try:
                out.append(
                    download_month(
                        year, month, dest_dir, session=session, force=force
                    )
                )
            except FileNotFoundError as exc:
                errors.append(str(exc))
    if errors:
        logger.warning("Download gaps (%d)", len(errors))
        for err in errors[:12]:
            logger.warning("  %s", err)
    return out


def _xlsx_product_label(row_vals: list[object]) -> Optional[str]:
    for idx in (2, 1, 0):
        if idx >= len(row_vals):
            continue
        val = row_vals[idx]
        if val is None:
            continue
        text = str(val).strip()
        if not text:
            continue
        if text.lower() in {"hauptprodukte:", "nebenprodukte:"}:
            return None
        if re.fullmatch(r"[a-zA-ZäöüÄÖÜß \-/(),.*]+", text) and len(text) > 2:
            return re.sub(r"\s+", " ", text)
    return None


def _detect_provisional_xlsx(path: Path) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Deckblatt" in wb.sheetnames:
            ws = wb["Deckblatt"]
            for row in ws.iter_rows(max_row=50, max_col=10, values_only=True):
                for cell in row:
                    if cell and "vorläufig" in str(cell).lower():
                        wb.close()
                        return True
        wb.close()
    except Exception:
        pass
    return False


def _detect_provisional_pdf(path: Path) -> bool:
    try:
        with pdfplumber.open(path) as pdf:
            text = _pdf_page_text(pdf.pages[0])
            return "vorläufig" in text.lower()
    except Exception:
        return False


def parse_xlsx_demand(path: Path, year: int, month: int) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Tab 6c" not in wb.sheetnames:
        raise KeyError(f"{path.name}: missing sheet 'Tab 6c'")
    ws = wb["Tab 6c"]
    date = pd.Timestamp(year=year, month=month, day=1)
    rows: list[dict] = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 12)]
        label = _xlsx_product_label(vals)
        if label is None:
            continue
        value = parse_german_number(vals[4] if len(vals) > 4 else None)
        if value is None:
            value = parse_german_number(vals[5] if len(vals) > 5 else None)
        if value is None:
            continue
        if label.lower() in {"mineralölprodukte", "zeitraum"}:
            continue
        rows.append(
            {
                "date": date,
                "metric_type": BAFA_DEMAND_METRIC,
                "product_native": label,
                "value": value,
                "unit": BAFA_UNIT_NATIVE,
                "source_file": path.name,
            }
        )
    return pd.DataFrame(rows)


def parse_xlsx_stocks(path: Path, year: int, month: int) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Tab 8" not in wb.sheetnames:
        raise KeyError(f"{path.name}: missing sheet 'Tab 8'")
    ws = wb["Tab 8"]
    date = pd.Timestamp(year=year, month=month, day=1)
    rows: list[dict] = []
    section = ""
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 12)]
        b = vals[1]
        if isinstance(b, str) and b.strip().lower().startswith("rohöl"):
            section = "Rohöl"
            continue
        if isinstance(b, str) and "mineralölprodukte" in b.strip().lower():
            section = "Mineralölprodukte"
            continue
        label_raw = vals[3]
        if label_raw is None:
            continue
        label = re.sub(r"\s+", " ", str(label_raw).strip())
        if not label or label.lower().startswith("bestandskategorie"):
            continue
        value = parse_german_number(vals[7] if len(vals) > 7 else None)
        if value is None:
            continue
        if section == "Rohöl" and label.lower() in {"deutsches", "ausländisches"}:
            product = f"Rohöl | {label}"
        elif label.lower() == "zusammen":
            product = f"{section} | Zusammen" if section else "Zusammen"
        else:
            product = label
        rows.append(
            {
                "date": date,
                "metric_type": BAFA_STOCKS_METRIC,
                "product_native": product,
                "value": value,
                "unit": BAFA_UNIT_NATIVE,
                "source_file": path.name,
            }
        )
    return pd.DataFrame(rows)


def parse_xlsx_bio(path: Path, year: int, month: int) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Tab 9" not in wb.sheetnames:
        raise KeyError(f"{path.name}: missing sheet 'Tab 9'")
    ws = wb["Tab 9"]
    date = pd.Timestamp(year=year, month=month, day=1)
    rows: list[dict] = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 12)]
        label = None
        for idx in (1, 0):
            if vals[idx] is None:
                continue
            text = str(vals[idx]).strip()
            if text and not text.startswith("Tabelle") and "Zeitraum" not in text:
                if text not in {"Biozusatzstoffe", "Ottokraftstoffe", "Dieselkraftstoff"}:
                    label = re.sub(r"\s+", " ", text)
                    break
        if label is None:
            continue
        value = parse_german_number(vals[3] if len(vals) > 3 else None)
        if value is None:
            continue
        rows.append(
            {
                "date": date,
                "metric_type": BAFA_BIO_METRIC,
                "product_native": label,
                "value": value,
                "unit": BAFA_UNIT_NATIVE,
                "source_file": path.name,
            }
        )
    return pd.DataFrame(rows)


_PDF_SECTION_MARKERS = {
    "demand": "Entwicklung der Inlandsablieferungen",
    "stocks": "Eigentumsendbestand",
    "bio": "Beimischung von Biozusatzstoffen",
}


def _pdf_page_text(page: pdfplumber.page.Page) -> str:
    text = page.extract_text() or ""
    if "(cid:" in text:
        text = decode_bafa_cid_text(text)
    return text


def _find_pdf_section_page(pdf: pdfplumber.PDF, needle: str) -> Optional[int]:
    body_hits: list[int] = []
    weak_hits: list[int] = []
    for i, page in enumerate(pdf.pages):
        text = _pdf_page_text(page)
        if needle not in text:
            continue
        if "Inhaltsverzeichnis" in text or text.count(". . .") > 5:
            continue
        if re.search(rf"\b[6-9]\.\s*{re.escape(needle)}", text):
            return i
        if i >= 2:
            body_hits.append(i)
        else:
            weak_hits.append(i)
    if body_hits:
        return body_hits[0]
    return weak_hits[0] if weak_hits else None


_LINE_VALUE_RE = re.compile(
    r"^(?P<label>.+?)\s+"
    r"(?P<v1>[+-]?[\d.]+(?:,\d+)?)\s+"
    r"(?P<v2>[+-]?[\d.]+(?:,\d+)?|[-–—])"
    r"(?:\s|$)"
)


def _parse_pdf_value_lines(
    text: str,
    *,
    date: pd.Timestamp,
    metric_type: str,
    source_file: str,
) -> list[dict]:
    rows: list[dict] = []
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line or line.startswith("Mengenangaben") or line.startswith("Zeitraum"):
            continue
        match = _LINE_VALUE_RE.match(line)
        if not match:
            continue
        label = match.group("label").strip(" .")
        if label.lower() in {
            "biozusatzstoffe",
            "berichtsmonat",
            "schmiermittel",
            "rohöl und produkte",
            "bestandskategorie",
        }:
            continue
        value = parse_german_number(match.group("v1"))
        if value is None:
            continue
        rows.append(
            {
                "date": date,
                "metric_type": metric_type,
                "product_native": label,
                "value": value,
                "unit": BAFA_UNIT_NATIVE,
                "source_file": source_file,
            }
        )
    return rows


def parse_pdf_sections(path: Path, year: int, month: int) -> pd.DataFrame:
    date = pd.Timestamp(year=year, month=month, day=1)
    frames: list[pd.DataFrame] = []
    with pdfplumber.open(path) as pdf:
        for key, needle in _PDF_SECTION_MARKERS.items():
            idx = _find_pdf_section_page(pdf, needle)
            if idx is None:
                continue
            text = _pdf_page_text(pdf.pages[idx])
            metric = {
                "demand": BAFA_DEMAND_METRIC,
                "stocks": BAFA_STOCKS_METRIC,
                "bio": BAFA_BIO_METRIC,
            }[key]
            rows = _parse_pdf_value_lines(
                text,
                date=date,
                metric_type=metric,
                source_file=path.name,
            )
            if rows:
                frames.append(pd.DataFrame(rows))
    if not frames:
        return pd.DataFrame(
            columns=[
                "date",
                "metric_type",
                "product_native",
                "value",
                "unit",
                "source_file",
            ]
        )
    return pd.concat(frames, ignore_index=True)


def parse_month_file(month_file: MonthFile) -> pd.DataFrame:
    """Parse one downloaded month into a source-native tidy frame."""
    if month_file.kind == "xlsx":
        parts = [
            parse_xlsx_demand(month_file.path, month_file.year, month_file.month),
            parse_xlsx_stocks(month_file.path, month_file.year, month_file.month),
            parse_xlsx_bio(month_file.path, month_file.year, month_file.month),
        ]
        df = pd.concat([p for p in parts if not p.empty], ignore_index=True)
        provisional = _detect_provisional_xlsx(month_file.path)
    else:
        df = parse_pdf_sections(month_file.path, month_file.year, month_file.month)
        provisional = _detect_provisional_pdf(month_file.path)

    if df.empty:
        return df

    df = df.copy()
    df["product_native"] = df["product_native"].map(normalize_product_native)
    df["is_provisional"] = provisional
    df = df[~df["product_native"].isin(_EXCLUDE_NATIVES)].copy()
    # Drop PDF line-wrap artefacts (e.g. "Hydrauliköle 350").
    junk = df["product_native"].str.contains(
        r"Hydrauliköle\s+\d|^\d", regex=True, na=False
    )
    df = df[~junk].copy()
    # Lubricant grade detail is not headline demand / stock.
    lubes_detail = {
        "Motorenöle",
        "Getriebeöle",
        "Basisöle",
        "Schmierfette",
        "Sonstige",
        "Maschinenöle",
        "Hydrauliköle",
        "IndustrieGetriebeöle",
        "Turbinen, Kompressoren und Elektroisolieröle",
    }
    if "metric_type" in df.columns:
        df = df[
            ~(
                df["product_native"].isin(lubes_detail)
                & df["metric_type"].isin({BAFA_DEMAND_METRIC, BAFA_STOCKS_METRIC})
            )
        ].copy()
    return df


# Exploration notebook alias.
parse_probe_file = parse_month_file


def finalize_bafa_frame(
    df: pd.DataFrame,
    *,
    updated_at: Optional[datetime] = None,
) -> pd.DataFrame:
    """Attach canonical country/source columns for the processor."""
    if df.empty:
        return pd.DataFrame(columns=CANONICAL_COLUMNS)

    out = df.copy()
    out["date"] = pd.to_datetime(out["date"])
    out["country"] = COUNTRY_CODE
    out["country_name"] = COUNTRY_NAME
    out["source"] = SOURCE_ID
    out["product"] = out["product_native"]
    out["unit"] = out.get("unit", BAFA_UNIT_NATIVE)
    if "is_provisional" not in out.columns:
        out["is_provisional"] = False
    out["updated_at"] = updated_at or datetime.now(UTC).replace(tzinfo=None)
    for col in CANONICAL_COLUMNS:
        if col not in out.columns:
            out[col] = pd.NA
    return out[CANONICAL_COLUMNS].sort_values(
        ["date", "metric_type", "product_native"], ignore_index=True
    )


def year_month_from_filename(path: Path) -> tuple[int, int]:
    """Parse ``moel_amtliche_daten_YYYY_MM.(xlsx|pdf)``."""
    match = re.search(r"moel_amtliche_daten_(\d{4})_(\d{1,2})\.", path.name, re.I)
    if not match:
        raise ValueError(f"Cannot parse year/month from {path.name}")
    return int(match.group(1)), int(match.group(2))


def list_local_month_files(raw_dir: Path) -> list[MonthFile]:
    raw_dir = Path(raw_dir)
    files: list[MonthFile] = []
    for path in sorted(raw_dir.glob("moel_amtliche_daten_*.*")):
        if path.suffix.lower() not in {".xlsx", ".pdf"}:
            continue
        try:
            year, month = year_month_from_filename(path)
        except ValueError:
            continue
        kind = "xlsx" if path.suffix.lower() == ".xlsx" else "pdf"
        files.append(MonthFile(path, kind, year, month))
    return files


def attach_value_kbd(df: pd.DataFrame) -> pd.DataFrame:
    """Add ``value_kbd`` for seasonality charts (tonnes → kbd)."""
    from analytics.units import convert_series

    out = df.copy()
    out["product_native"] = out["product_native"].map(normalize_product_native)
    out["product_kind"] = out["product_native"].map(UNITS_KIND)
    out = out.dropna(subset=["product_kind", "value", "date"]).copy()
    out["value_kbd"] = convert_series(
        out["value"] / 1000.0,
        from_unit="kt",
        to_unit="kbd",
        product_kind=out["product_kind"],
        date=out["date"],
    )
    return out


def seasonality_chart_inputs(
    long: pd.DataFrame,
    *,
    metric_type: str = BAFA_DEMAND_METRIC,
    products: Optional[tuple[str, ...]] = None,
) -> tuple[pd.DataFrame, str, list[str], dict[str, str], str]:
    if products is None:
        products = (
            SEASONALITY_BIO_PRODUCTS
            if metric_type == BAFA_BIO_METRIC
            else SEASONALITY_DEMAND_PRODUCTS
        )
    base = long[long["metric_type"] == metric_type].copy()
    demand_kbd = attach_value_kbd(base)
    present = [p for p in products if p in set(demand_kbd["product_native"])]
    df = demand_kbd[demand_kbd["product_native"].isin(present)].copy()
    labels = {p: DISPLAY_LABELS.get(p, p) for p in present}
    suffix = "bio blends" if metric_type == BAFA_BIO_METRIC else "native products"
    return df, "product_native", present, labels, suffix


def bafa_series_for_jodi(
    demand: pd.DataFrame,
    series_key: str,
    *,
    value_col: str = "value_kbd",
) -> pd.DataFrame:
    return sum_natives_series_for_jodi(
        demand,
        series_key,
        jodi_compare=JODI_COMPARE_SERIES,
        value_col=value_col,
    )


def describe_file(path: Path) -> dict:
    """Format fingerprint helper (exploration notebook)."""
    path = Path(path)
    info: dict = {
        "file": path.name,
        "bytes": path.stat().st_size,
        "suffix": path.suffix.lower(),
    }
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True)
        info["sheets"] = wb.sheetnames
        info["has_tab_6c"] = "Tab 6c" in wb.sheetnames
        info["has_tab_8"] = "Tab 8" in wb.sheetnames
        info["has_tab_9"] = "Tab 9" in wb.sheetnames
        wb.close()
        return info
    with pdfplumber.open(path) as pdf:
        info["pages"] = len(pdf.pages)
        sample = pdf.pages[0].extract_text() or ""
        info["cid_encoded"] = "(cid:" in sample
        if info["cid_encoded"]:
            info["decoded_title_sample"] = decode_bafa_cid_text(sample)[:120]
        else:
            info["title_sample"] = sample[:120]
    return info
